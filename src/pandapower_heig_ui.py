# Import package
import os
import logging
import coloredlogs
import traceback
from copy import deepcopy
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, time, timedelta
import pandapower as pp
import pandapower.plotting.plotly as pplotly
import pandapower.control as control
import pandapower.timeseries as timeseries
from pandapower.timeseries.data_sources.frame_data import DFData
import plotly.graph_objects as go
from plotly.graph_objs import Figure, Layout
from plotly.graph_objs.layout import XAxis, YAxis

log = logging.getLogger(__name__)
coloredlogs.install(level="INFO")

def parse_grid_from_xlsx(file_path: str) -> pp.pandapowerNet:
 
    # Columns which have to be set by default when None value is founded
    default_values: dict = {
        "in_service": True, "g_us_per_km": 0.0, "g0_us_per_km": 0.0, "r0_ohm_per_km": 0.0, "x0_ohm_per_km": 0.0,
        "c0_nf_per_km": 0.0, "parallel": 1, "df": 1.0, "p_mw": 0.0, "q_mvar": 0.0, "const_z_percent": 0.0,
        "const_i_percent": 0.0, "scaling": 1.0, "vk0_percent": 0.0, "vkr0_percent": 0.0, "mag0_percent": 0.0,
        "mag0_rx": 0.0, "si0_hv_partial": 0.0, "shift_degree": 0.0, "tap_step_percent": 1.0, "tap_phase_shifter": False,
        "tap_step_degree": 0.0, "vm_pu": 1.0, "va_degree": 0.0, "slack_weight": 1.0, "tap_pos": 0, "tap_neutral": 0,
        "tap_min": 0, "tap_max": 1, "profile_mapping": -1, "current_source": True
    }
    # Columns which as to be converted to another type
    int_column: list[str] = ["bus", "parallel", "from_bus", "to_bus", "element", "hv_bus", "lv_bus", "tap_pos",
                             "tap_neutral", "tap_min", "tap_max", "profile_mapping"]
    bool_column: list[str] = ["current_source", "in_service"]
    # Create empty network
    net: pp.pandapowerNet = pp.create_empty_network()
    # Open Excel file and iterate over every existing sheet table
    for eq_name in openpyxl.load_workbook(file_path).sheetnames:
        # Create a dataFrame form an Excel sheet table
        data_df: pd.DataFrame = pd.read_excel(file_path, sheet_name=eq_name).drop(columns="idx").dropna(how="all")
        if not data_df.empty:
            # Fill null values using default_values dictionary
            data_df.fillna(value=default_values, inplace=True)
            # Fill bus and load types
            if eq_name == "bus":
                data_df.fillna(value={"type": "b"}, inplace=True)
            elif eq_name == "load":
                data_df.fillna(value={"type": "wye"}, inplace=True)
            elif eq_name == "line_geodata":
                # Create list of coordinates from string
                data_df["coords"] = data_df.coords.apply(
                    lambda x: list(map(lambda y: [float(z) for z in y.split(",")],
                                       x.replace("[[", "").replace("]]", "").split("], ["))))
            # Change needed columns type from float to int64
            col: list[str] = list(set(data_df.columns).intersection(int_column))
            data_df[col] = data_df[col].astype('int64')
            # Change needed columns type from float to bool
            col: list[str] = list(set(data_df.columns).intersection(bool_column))
            data_df[col] = data_df[col].astype(bool)
            # Replace np.nan to None
            data_df = data_df.replace(np.nan, None)
            # Create pandapower network
            net[eq_name] = data_df
    return net


def load_power_profile(file_path: str) -> dict[str, dict[str, pd.DataFrame]]:
    """

    Args:
        file_path:

    Returns:

    """
    # Initialize power profiles dictionaries
    power_profile: dict[str, pd.DataFrame] = dict()
    results: dict[str, dict[str, pd.DataFrame]] = dict()
    # Initialize times which will be used to interpolate power profiles
    period = timedelta(days=1)
    end_time = time()
    start_time = time(23, 59, 59)
    # Load every Excel sheet names
    for eq_name in openpyxl.load_workbook(file_path).sheetnames:
        # Create a dataFrame form an Excel sheet table
        data_df = pd.read_excel(
            file_path, sheet_name=eq_name, header=[0, 1], index_col=0
            ).dropna(how="all", axis=1).dropna(how="all", axis=0)
        # If there is no data within the Excel sheet don't save the dataFrame
        if not data_df.empty:
            # Save DataFrame in power profile dictionary 
            power_profile[eq_name] = data_df
            # Find the most suitable common period, start time and end time
            date_time = pd.Series(data_df.index).apply(lambda x: datetime.combine(datetime.today(), x))
            period = min(period, date_time.diff().min())
            start_time = min(start_time, data_df.index[0])
            end_time = max(end_time, data_df.index[-1])

    # Create common datetime index for every power profiles
    datetime_index = pd.Series(
        pd.date_range(
            start= datetime.combine(datetime.today(),start_time),
            end= datetime.combine(datetime.today(),end_time),
            freq=period
        ).to_pydatetime()).apply(lambda x: x.time())
    # Create a dataframe with common time and its corresponding timestamp number
    datetime_inter = pd.DataFrame(
        range(datetime_index.shape[0]), index=datetime_index,
        columns=["timestamp"]
    )
    for eq_name, eq_power in power_profile.items():
        # Create multiindex columns from power profile
        columns = pd.MultiIndex.from_tuples(
            eq_power.columns, names=["profile", "power"]
        )
        # Concatenate power profile with created common datetime index
        data_df  = pd.concat([datetime_inter, eq_power], axis=1).sort_index()
        # Fill timestamps unmatched using closest neighbors (forward and backward in order to fill every null)
        data_df["timestamp"] = data_df["timestamp"].fillna(method='ffill').fillna(method='bfill')
        # Group timestamp to keep only the first founded non-null value and then fill null values using
        # first order interpolation
        data_df =  data_df\
            .groupby("timestamp").first()\
            .interpolate().fillna(method="bfill").fillna(method="ffill")\
        # Create interpolated dataframe with good index and apply into results dictionary
        actual_profile = pd.DataFrame(
            data_df.values, index=datetime_inter.index, columns=columns
        )
        results[eq_name] = dict()
        if "P [MW]" in eq_power.columns.levels[1]:
            results[eq_name]["p_mw"] = actual_profile.xs("P [MW]", level=1,axis=1).dropna(how="all", axis=1)
        if "Q [MVAR]" in eq_power.columns.levels[1]:
            results[eq_name]["q_mvar"] =  actual_profile.xs("Q [MVAR]", level=1,axis=1).dropna(how="all", axis=1)
    return results


def apply_power_profile(net: pp.pandapowerNet, equipment: str, power_profiles: dict[str, pd.DataFrame]):

    # Check if pandapower network already have controllers for the given equipment and delete them
    if not net.controller.empty:
        old_controller_index = net.controller[net.controller.object.apply(lambda x: x.element) == equipment].index
        net.controller = net.controller.drop(old_controller_index).reset_index(drop=True)
    # Create a dictionary from profile_mapping column
    profile_mapping: dict = net[equipment]\
        .reset_index()\
        .groupby("profile_mapping")["index"]\
        .apply(list).to_dict()
    for variable, profile in power_profiles.items():
        if profile is not None:
            #
            if "time_index" in net.keys():
                if not (net["time_index"] == pd.DataFrame(profile.index, columns=["Time"])).all().all():
                    log.error("Simulation Profiles have not the same timestamps")
            else:
                net["time_index"] = pd.DataFrame(profile.index, columns=["Time"])
            if profile_mapping:
                mapped_profile = pd.DataFrame(index=profile.index)
                for profile_id, eq_list in profile_mapping.items():
                    if profile_id in profile.columns:
                        mapped_profile[eq_list] = list(zip(*[list(profile[profile_id].values)] * len(eq_list)))
            else:
                mapped_profile = profile[net[equipment].index.intersection(profile.columns)]
            unmapped_equipment = net[equipment].index.difference(mapped_profile.columns)
            if len(unmapped_equipment) != 0:
                unmapped_eq_names = ", ".join(net[equipment].loc[unmapped_equipment, "name"].values)
                log.warning("{} equipments have no {} profiles".format(unmapped_eq_names, variable))
            mapped_profile.reset_index(drop=True, inplace=True)

            control.ConstControl(net, element=equipment, element_index=mapped_profile.columns,
                                 variable=variable, data_source=DFData(mapped_profile),
                                 profile_name=mapped_profile.columns)

def create_output_writer(net: pp.pandapowerNet, add_results: list[str] = None):

    
    results =  add_results + ["res_trafo.loading_percent"] if add_results else ["res_trafo.loading_percent"]
    ow = timeseries.OutputWriter(net)
               
    for result in results:
        ow.log_variable(*result.split("."))

def run_time_simulation(
        net: pp.pandapowerNet, output_filename: str = None, folder: str= r"output"
) -> dict[str, pd.DataFrame]:

    timeseries.run_timeseries(net, time_steps=range(net["time_index"].shape[0]), verbose=False)
    results_df: dict[str, pd.DataFrame] = dict()
    for key, result in net.output_writer.at[0, "object"].output.items():
        if key != "Parameters":
            results_df[key] = result.set_index(net["time_index"].Time)
            eq = key.replace("res_", "").split(".")[0]
            name_mapping = dict(zip(net[eq].index, net[eq].name))
            results_df[key] = results_df[key].rename(columns=name_mapping)

    if output_filename:
        output_file_path = os.path.join(folder, output_filename + ".xlsx")
        try:
            if not os.path.exists(folder):
                os.makedirs(folder)
            writer: pd.ExcelWriter = pd.ExcelWriter(output_file_path)
            for keys in results_df.keys():
                results_df[keys].to_excel(writer, sheet_name=keys, index_label="Time")
            # writer.save()
        except(Exception,):
            log.error("Error in {} function:\n".format(traceback.extract_stack()[-1].name) +
                        "Impossible to save simulation results in {} file\n".format(output_file_path) +
                        "Check output file path or if file is already open in your computer")
    return results_df

def plot_net_by_zone(net: pp.pandapowerNet, plot_title: str = None, filename: str = None, folder: str = "plot",
                     vertical_tree: bool = True, line_width: int = 3,
                     bus_size: int = 20, **kwargs):
    net_copy = deepcopy(net)
    colors = ['b', 'g', 'r', 'c', 'm', 'k', 'w']
    # Generate bus geodata if needed
    if net_copy.bus_geodata.empty:
        pplotly.create_generic_coordinates(net=net_copy, overwrite=True)
        # Rotate figure if wanted
        if vertical_tree:
            net_copy.bus_geodata[["x", "y"]] = net_copy.bus_geodata[["y", "x"]]
            net_copy.bus_geodata["y"] *=-1

    traces = []
    # create lines trace
    traces += pplotly.create_line_trace(net_copy, net_copy.line.index, color='k', width=line_width,
                                             trace_name="Lines")
    # Create transfo trace
    trafo_info = pd.Series(
        index=net_copy.trafo.index,
        data=net_copy.trafo.name + '<br>' + net_copy.trafo.sn_mva.astype(str) + ' MVA'
    )
    traces += pplotly.create_trafo_trace(
        net_copy, trafos=net_copy.trafo.index, width=int(2.5*line_width), trace_name="Transformers",
        infofunc = trafo_info
    )
    # create ext_grid trace
    traces += pplotly.create_bus_trace(
        net_copy, buses=net_copy.ext_grid.bus.values, patch_type="square", size=int(1.5*bus_size), color='y',
        infofunc=pd.Series(index=net_copy.ext_grid.bus.values, data=net_copy.ext_grid.name),
        trace_name="Ext grid"
    )
    # create bus trace
    bus_info = pd.Series(
        index=net_copy.bus.index,
        data=net_copy.bus.name + '<br>' + net_copy.bus.vn_kv.astype(str) + ' kV')
    for i, zone in enumerate(net_copy.bus.zone.unique()):
        bus_index = net_copy.bus[net_copy.bus.zone == zone].index
        traces += pplotly.create_bus_trace(
            net_copy, buses=bus_index, size=bus_size, color=colors[i],
            infofunc=bus_info.loc[bus_index],
            trace_name="Zone {}".format(zone)
        )  # create buses
    fig = _draw_traces(traces=traces, showlegend=True)
    save_fig(fig=fig, filename=filename, folder=folder, plot_title=plot_title, **kwargs)


def plot_net_simple_powerflow_result(
        net: pp.pandapowerNet, filename: str = None, folder: str="output", plot_title: str=None,
        line_width: int = 3, trafo_width: int =7, bus_size: int = 20, voltage_cmap: str = "jet",
        loading_cmap: str =  "jet", voltage_range: tuple[float] = (0.85, 1.15),
        loading_range: tuple[int] = (0, 100), width: int= 770, **kwargs
        ):
    """Plot network scheme in plotly figure with powerflow results displayed

    INPUT:
        **net** (pandapower.pandapowerNet): pandaPower network object with powerflow results stored. Powerflow
        simulations are described in
        `pandaPower powerflow docs <https://pandapower.readthedocs.io/en/v2.0.1/powerflow.html>`_.

    OPTIONAL:
        **filename** (str, None): File name under which the plotly figure will be stored (in png format).
        If this parameter is not filled, the function will only display the figure without saving it.

        **folder** (str, "output"): Folder name where the plotly figure will be stored.

        **plot_title** (str, None): Title displayed on the top of the figure.

        **line_width** (int, 3): Transmission lines drawing width.

        **trafo_width** (int, 7): Transformers drawing width.

        **bus_size** (int, 20): Bus drawing size.

        **voltage_cmap** (str, "jet"): name of a plotly Continuous Color Scales used to display voltage results (Greys,
        YlGnBu, Greens, YlOrRd, Bluered, RdBu, Reds, Blues, Picnic, Rainbow, Portland, Jet, Hot, Blackbody, Earth,
        Electric, Viridis). Further explanation are founded in
        `plotly docs <https://plotly.com/python/builtin-colorscales/?_ga=2.67899217.1309821379.1693317794-265230606.1688628396>`_.

        **loading_cmap** (str, "jet"): name of a plotly Continuous Color Scales used to display line loading results.

        **voltage_range** (tuple[float], (0.85, 1.15)): Voltage range used by the plotly Continuous Color Scales
        (voltage units are in p.u.).

        **width** (int, 770): Plotly figure width.

        **loading_range** (tuple[float], (0, 100)): Voltage range used by the plotly Continuous Color Scales
        (loading units are in %).

        **kwargs**: Every parameter found in save_fig function could also be added if needed

    """
    net_copy = deepcopy(net)
    traces = []
    # Generate bus geodata if needed
    if net_copy.bus_geodata.empty:
        pplotly.create_generic_coordinates(net=net_copy, overwrite=True)

        net_copy.bus_geodata[["x", "y"]] = net_copy.bus_geodata[["y", "x"]]
        net_copy.bus_geodata["y"] *=-1

    trafo_info = pd.Series(
        index=net_copy.line.index,
        data=net_copy.trafo.name + '<br>' + "Loading: " +
             net_copy.res_trafo.loading_percent.apply(lambda x: str(round(x, 1))) + ' %'
    )
    traces += pplotly.create_trafo_trace(
        net_copy, trafos=net_copy.trafo.index, width=trafo_width, trace_name="Transformers",
        cmap=loading_cmap, cmin=0, cmax=100,infofunc=trafo_info,
    )

    line_info = pd.Series(
        index=net_copy.line.index,
        data=net_copy.line.name + '<br>' + "Loading: " +
             net_copy.res_line.loading_percent.apply(lambda x: str(round(x, 1))) + ' %'
    )
    traces += pplotly.create_line_trace(
        net_copy, lines=net_copy.line.index, cmap=loading_cmap, width=line_width, cmin=loading_range[0],
        cmax=loading_range[1], cpos=1.15, infofunc=line_info, cbar_title='Equipment loading [%]')
    traces += pplotly.create_bus_trace(
        net_copy, buses=net_copy.ext_grid.bus.values, patch_type="square", size=int(1.5*bus_size), color='y',
        infofunc=pd.Series(index=net_copy.ext_grid.bus.values, data=net_copy.ext_grid.name),
        trace_name="Ext grid"
    )
    bus_info =pd.Series(
        index=net_copy.bus.index,
        data=net_copy.bus.name + '<br>' + "V: " + net_copy.res_bus.vm_pu.apply(lambda x: str(round(x, 3))) + ' pu'
    )
    traces += pplotly.create_bus_trace(
        net_copy, buses=net_copy.bus.index, cmap=voltage_cmap, size=bus_size, cmin=voltage_range[0],
        cmax=voltage_range[1], infofunc=bus_info, cbar_title='Bus voltage [pu]'
    )
    fig = _draw_traces(traces, showlegend=False)
    save_fig(fig=fig, filename=filename, folder=folder, plot_title=plot_title, width=width, **kwargs)


def plot_net_short_circuit_result(
        net: pp.pandapowerNet, filename: str=None, folder: str="plot", plot_title:str=None,
        line_width: int = 3, trafo_width: int =7, bus_size: int = 20,
        cmap: str = "jet", **kwargs
    ):
    """Plot network scheme in plotly figure with short-circuit results displayed

    INPUT:
        **net** (pandapower.pandapowerNet): pandaPower network object with short-circuit results stored. Short-circuit
        simulations are described in
        `pandaPower short-circuit docs <https://pandapower.readthedocs.io/en/v2.0.1/shortcircuit.html>`_.

    OPTIONAL:
        **filename** (str, None): File name under which the plotly figure will be stored (in png format).
        If this parameter is not filled, the function will only display the figure without saving it.

        **folder** (str, "plot"): Folder name where the plotly figure will be stored.

        **plot_title** (str, None): Title displayed on the top of the figure

        **line_width** (int, 3): Transmission lines drawing width.

        **trafo_width** (int, 7): Transformers drawing width.

        **bus_size** (int, 20): Bus drawing size.

        **cmap** (str, "jet"): name of a plotly Continuous Color Scales used to display short-circuit results (Greys,
        YlGnBu, Greens, YlOrRd, Bluered, RdBu, Reds, Blues, Picnic, Rainbow, Portland, Jet, Hot, Blackbody, Earth,
        Electric, Viridis). Further explanation are founded in
        `plotly docs <https://plotly.com/python/builtin-colorscales/?_ga=2.67899217.1309821379.1693317794-265230606.1688628396>`_.

        **kwargs**: Every parameter found in save_fig function could also be added if needed.
    """
    net_copy = deepcopy(net)
    traces = []
    # Generate bus geodata if needed
    if net_copy.bus_geodata.empty:
        pplotly.create_generic_coordinates(net=net_copy, overwrite=True)
        # Rotate figure
        net_copy.bus_geodata[["x", "y"]] = net_copy.bus_geodata[["y", "x"]]
        net_copy.bus_geodata["y"] *=-1


    traces += pplotly.create_trafo_trace(
        net_copy, trafos=net_copy.trafo.index, width=trafo_width, trace_name="Transformers"
    )

    traces += pplotly.create_line_trace(net_copy, net_copy.line.index, color='k', width=line_width,
                                             trace_name="Lines")

    bus_info =pd.Series(
        index=net_copy.bus.index,
        data=net_copy.bus.name + '<br>' + "ikss: " + net_copy.res_bus_sc.ikss_ka.apply(lambda x: str(round(x, 3))) +
             ' kA<br>ip: ' + net_copy.res_bus_sc.ip_ka.apply(lambda x: str(round(x, 3))) + ' kA'
    )
    traces += pplotly.create_bus_trace(
        net_copy, buses=net_copy.ext_grid.bus.values, patch_type="square", size=int(1.5*bus_size), color='y',
        infofunc=pd.Series(index=net_copy.ext_grid.bus.values, data=net_copy.ext_grid.name),
        trace_name="Ext grid"
    )
    traces += pplotly.create_bus_trace(
        net_copy, buses=net_copy.bus.index, cmap=cmap, size=bus_size, cmin=0,
        cmax=net_copy.res_bus_sc.ikss_ka.max(), cmap_vals= list(net_copy.res_bus_sc.ikss_ka),
        infofunc=bus_info, cbar_title='Bus short circuit current [kA]'
    )

    fig = _draw_traces(traces, showlegend=False)
    save_fig(fig=fig, filename=filename, folder=folder, plot_title=plot_title, **kwargs)


def plot_net_time_simulation_result(
        net: pp.pandapowerNet, plot_time: time, filename: str = None, folder: str="plot", **kwargs
    ):
    """Plot network scheme in plotly figure with powerflow results from one chosen timestep displayed

     INPUT:
         **net** (pandapower.pandapowerNet): pandaPower network object with time simulation powerflow results stored.
          Powerflow simulations are described in
         `pandaPower time simulation powerflow docs <https://pandapower.readthedocs.io/en/v2.0.1/powerflow.html>`_.
         **plot_time** (datetime.time): Timestep used to plot results

     OPTIONAL:
         **filename** (str, None): File name under which the plotly figure will be stored (in png format).
         If this parameter is not filled, the function will only display the figure without saving it.

         **folder** (str, "plot"): Folder name where the plotly figure will be stored.

         **plot_title** (str, None): Title displayed on the top of the figure.

         **kwargs**: Every parameter found in save_fig and
         plot_net_simple_powerflow_result functions could also be added if needed.
    """
    net_copy = deepcopy(net)
    plot_title = "Powerflow results at " + plot_time.strftime("%Hh%M")
    if "output_writer" in net_copy.keys():
        ow = net_copy.output_writer.at[0, "object"].output
        idx = net_copy.time_index.Time.searchsorted(plot_time)
        net_copy.res_bus.vm_pu = list(ow["res_bus.vm_pu"].iloc[idx])
        net_copy.res_line.loading_percent = list(ow["res_line.loading_percent"].iloc[idx])
        net_copy.res_trafo.loading_percent = list(ow["res_trafo.loading_percent"].iloc[idx])
        plot_net_simple_powerflow_result(
            net=net_copy, plot_title=plot_title, filename=filename, folder=folder, **kwargs
        )
    else:
        log.error("PandaPower network does not contain time simulation powerflow results")


def plot_timeseries_result(
        data_df: pd.DataFrame, ylabel:str, plot_title: str = None, filename: str = None, folder: str="plot",
        **kwargs
    ):
    """

    INPUT:
        **data_df** (pandas.Dataframe): DataFrame which contains timeseries simulation results which will be plotted

        **ylabel** (str): Y-axis label

    OPTIONAL:
        **plot_title** (str, None):  Title displayed on the top of the figure.

        **filename** (str, None): File name under which the plotly figure will be stored (in png format).
         If this parameter is not filled, the function will only display the figure without saving it.

        **folder** (str, "plot"): Folder name where the plotly figure will be stored.

        **\*\*kwargs**: Every parameter found in save_fig function could also be added if needed.
    """
    fig = Figure()
    for col in data_df.columns:
        hovertemplate = _bold(col + '<br>' + ylabel + ': %{y:.2f}<br>Time: %{text}') + '<extra></extra>'
        fig.add_trace(go.Scatter(
            x=list(range(data_df.shape[0])), y=list(data_df[col]), mode='lines',
            name=_bold(col), hovertemplate=hovertemplate,
            text=_time_to_str(data_df.index)
        ))
    x_ticks = list(range(data_df.shape[0]))[::int(data_df.shape[0] / 6)]
    x_ticks_label = _time_to_str(data_df.index)[::int(data_df.shape[0] / 6)]
    save_fig(
        fig=fig, filename=filename, folder=folder, xlabel= "Time", ylabel=ylabel, plot_title=plot_title,
        x_ticks=(x_ticks,x_ticks_label), show_grid=True, **kwargs
    )


def _draw_traces(traces: list, showlegend: bool) -> Figure:
    """
    Intern function used to create a plotly figure from a list of trace created by pandasPower using `create_bus_trace`,
    `create_line_trace` and `create_trafo_trace`
     `functions <https://pandapower.readthedocs.io/en/v2.0.1/powerflow.html>`_.
    to `PLOTLY <https://plot.ly/python/>`

    INPUT:
        **traces** - list of dicts which correspond to plotly traces
        generated using: `create_bus_trace`, `create_line_trace`, 
        `create_trafo_trace`

    OUTPUT:
        **figure** (graph_objs._figure.Figure) figure object

    """

    fig = Figure(data=traces,  # edge_trace
                 layout=Layout(
                     showlegend=showlegend,
                     hovermode='closest',
                     margin=dict(b=5, l=5, r=5, t=5),
                     xaxis=XAxis(showgrid=False, zeroline=False, 
                                 showticklabels=False),
                     yaxis=YAxis(showgrid=False, zeroline=False, 
                                 showticklabels=False),
                     legend={'itemsizing': 'constant'},
                     ),
                )
    return fig


def save_fig(
        fig: Figure, filename: str = None, folder: str = "plot", plot_title: str = None,
        xlabel: str = None, ylabel: str = None, 
        x_ticks: tuple[list, list] | None = None,
        y_ticks: tuple[list, list] | None = None,
        width: int = 680, height: int = 400, title_x: float = 0.5, title_y: float = 0.97,
        legend_size: int = 12, tick_size: int = 12, axis_title_size: int = 12,
        title_size: int = 15, show_grid: bool = False, show_fig: bool = True
    ):
    """Intern function used to format plotly layout, save and display figure

    INPUT:
        **fig** (graph_objs._figure.Figure):

    OPTIONAL:
        **filename** (str, None): File name under which the plotly figure will be stored (in png format).
        If this parameter is not filled, the function will only display the figure without saving it.

        **folder** (str, "plot"): Folder name where the plotly figure will be stored.

        **plot_title** (str, None): Title displayed on the top of the figure.

        **xlabel** (str, None): X-axis label.

        **ylabel** (str, None): Y-axis label.

        **x_ticks** (tuple[list, list], None): A pair of tick values and tick label for displaying x-axis.

        **y_ticks** (tuple[list, list], None): A pair of tick values and tick label for displaying y-axis.

        **width** (int, 680): Figure width (in pixels).

        **height** (int, 400): Figure height (in pixels).

        **title_x** (float, 0.5): Figure title x-axis location.

        **title_y** (float, 0.97): Figure title Y-axis location.

        **legend_size** (int, 12): Legends characters size.

        **tick_size (int, 12): Ticks characters size.

        **axis_title_size** (int, 12): Axis title characters size.

        **title_size** (int, 15): Figure title characters size.

        **show_grid** (bool, False): If it is True, add grid to the figure.

        **show_fig** (bool, True): If it is True, display figure in Jupyter notebook (usage with a .py file is not
        implemented).

    """
    fig.update_layout(
        font={"size": title_size}, xaxis_title=_bold(xlabel), yaxis_title=_bold(ylabel),
        xaxis=dict(tickfont=dict(size=tick_size, family="Arial Black"), zeroline=False),
        yaxis=dict(tickfont=dict(size=tick_size, family="Arial Black"), zeroline=False),
        paper_bgcolor='white', plot_bgcolor='white', width = width, height = height,
        xaxis_title_font = {"size": axis_title_size}, yaxis_title_font = {"size": axis_title_size},
        margin = dict(t=30), legend = {'font': {'size': legend_size}}
    )
    if x_ticks:
        fig.update_layout(
            xaxis=dict(tickmode = 'array', tickvals = x_ticks[0], ticktext = x_ticks[1])
        )
    if y_ticks:
        fig.update_layout(
            yaxis=dict(tickmode = 'array', tickvals = y_ticks[0], ticktext = y_ticks[1])
        )
    if show_grid:
        fig.update_layout(
            xaxis=dict(showline=True, showgrid=True, gridwidth= 1, linewidth=2, linecolor='black', gridcolor='black'),
            yaxis=dict(showline=True, showgrid=True, gridwidth= 1, linewidth=2, linecolor='black', gridcolor='black')
        )
    if plot_title:
        fig.update_layout(
            title={'text': _bold(plot_title), 'y': title_y, 'x': title_x, 'xanchor': 'center', 'yanchor': 'top'},
        )
    if filename:
        if not os.path.exists(folder):
            os.makedirs(folder)
        file_path = os.path.join(folder, filename)
        # plotly.offline.plot(fig, filename=file_path + ".html", auto_open=False)
        fig.write_image(file_path + ".png")
    if show_fig:
        fig.show()

def _bold(string: str | None) -> str | None:
    """Intern function to apply bold style to strings

    INPUT:
    **string** (str | None): input string.

    OUTPUT:
    **bold_string** (str | None): input string with bold style applied.

    """
    if isinstance(string, str):
        return '<b>' + string + '<b>'
    else:
        return None

def _time_to_str(time_stamps: [time | datetime | list | tuple | pd.Index | pd.Series]) -> str | list[str] | None:
    """
    Intern function used to convert datetime objects in wanted string format (i.e. "12H00")
    INPUT:
    **time_stamps** (datetime.time | datetime.datetime | list | tuple | pandas.Index | pandas.Series): Input datetime.

    OUTPUT:
    **str_time_stamps** (str | list[str] | None): Datetime objects in wanted string format

    """
    if isinstance(time_stamps, (time, datetime)):
        return time_stamps.strftime("%Hh%M")
    elif isinstance(time_stamps, (list, tuple, pd.Index, pd.Series)):
        return list(map(lambda x: x.strftime("%Hh%M"), time_stamps))
    else:
        return None